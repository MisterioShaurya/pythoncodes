from openpyxl import load_workbook

def display_excel(filename):
    workbook = load_workbook(filename)
    for sheet_name in workbook.sheetnames:
        print(f"Sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            print(row)

if __name__ == "__main__":
    filename = input("Enter the path to the Excel file: ")
    display_excel(filename)
